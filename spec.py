import asyncio
import asyncpg
from aiogram import Bot, Dispatcher, F
from aiogram.types import Message, ReplyKeyboardRemove, ReplyKeyboardMarkup, KeyboardButton, Location
from aiogram.filters import CommandStart
from aiogram.enums.parse_mode import ParseMode
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.client.default import DefaultBotProperties
from aiogram.types.input_file import FSInputFile
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import Font, Alignment
from aiogram.types import BufferedInputFile
from openpyxl.utils import get_column_letter
import logging
import json
import os
import io

# --- Logging sozlamalari ---
logging.basicConfig(level=logging.ERROR, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# --- Config ---
BOT_TOKEN = '7731234849:AAFoVPyPLSiuAiXaOaw8q1_A3bjDE51NhSI'
DB_CONFIG = {
    'user': 'postgres',
    'password': 'saman07',
    'database': 'bp',
    'host': 'localhost',
    'port': 5432
}
GROUP_CHAT_ID = "-1002865807540"

# Dorilar nomlarini mapping qilish uchun lug'at
MEDICINE_MAPPING = {
    "БРЕЙНЦИТ": "BREYNCIT",
    "ВИРАГЕМ": "VIRAGEM",
    "ИЗОЛИТ": "IZOLIT",
    "КАЛЕРОН": "KALERON",
    "КАРНИКИД": "KARNIKID",
    "РЕЛАФЛОР": "RELAFLOR",
    "СИНИМАГ": "SINIMAG",
    "СОНАЙТ": "SONAYT",
    "ЗИНКИД": "ZINKID",
    "РЕСЛИП": "RESLIP",
    "ЭЛВИКИД": "ELVIKID",
    "МАГНИЙ Б6": "MAGNIY B",
    "КАЛЦИЙ+Д3": "KALTSIY+D3",
    "АРТРОЗИТ": "ARTROZIT"
}

REVERSE_MEDICINE_MAPPING = {v: k for k, v in MEDICINE_MAPPING.items()}

# Load doriops.json
with open('C:\\Users\\user\\Desktop\\BOTS\\doriops.json', 'r', encoding='utf-8') as f:
    DORIOPS = json.load(f)

# --- FSM state ---
class StartWorkState(StatesGroup):
    waiting_for_phone = State()
    waiting_for_location = State()
    waiting_for_video = State()

class SpecState(StatesGroup):
    selecting_search_type = State()
    waiting_for_inn = State()
    waiting_for_name = State()
    selecting_from_list = State()
    confirming_apteka = State()

class OrderState(StatesGroup):
    selecting_medicine = State()
    entering_quantity = State()
    reviewing_order = State()

class EndWorkState(StatesGroup):
    waiting_for_location = State()
    waiting_for_video = State()

# --- Init ---
bot = Bot(token=BOT_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher(storage=MemoryStorage())

# --- Database functions ---
async def get_medicine_details(medicine_name: str, pool: asyncpg.Pool = None) -> tuple[float | None, int | None]:
    query = """SELECT price, upk FROM dori WHERE l_name = $1"""
    try:
        if pool:
            async with pool.acquire() as conn:
                row = await conn.fetchrow(query, medicine_name)
        else:
            async with asyncpg.connect(**DB_CONFIG) as conn:
                row = await conn.fetchrow(query, medicine_name)
        if row:
            return row['price'], row['upk']
        else:
            logger.warning(f"Dori topilmadi: {medicine_name}")
            return None, None
    except Exception as e:
        logger.error(f"DB xatosi dori uchun {medicine_name}: {e}")
        return None, None

async def get_total_price(order, pool: asyncpg.Pool = None) -> tuple[float, float]:
    total = 0
    discountable_total = 0
    non_discountable_medicines = {'MAGNIY B', 'KALTSIY+D3', 'ARTROZIT'}
    for medicine, qty in order.items():
        try:
            price, _ = await get_medicine_details(MEDICINE_MAPPING.get(medicine, medicine), pool)
            if price:
                total += price * qty
                if medicine not in non_discountable_medicines:
                    discountable_total += price * qty
                logger.info(f"{medicine}: {qty} ta x {price} = {price*qty} so'm")
            else:
                logger.warning(f"{medicine} uchun narx mavjud emas")
        except Exception as e:
            logger.error(f"Price fetch error for {medicine}: {e}")
    logger.info(f"Umumiy narx: {total} so'm, Chegirma qo'llaniladigan narx: {discountable_total} so'm")
    return total, discountable_total

async def send_to_group(bot: Bot, order: dict, apteka_info: dict, total_price: float, discounted_price: float, is_full_payment: bool, user_info: dict):
    try:
        message_text = f"<b>Yangi buyurtma!</b>\n\n"
        message_text += f"<b>Dorixona:</b> {apteka_info['firma']}\n"
        message_text += f"<b>INN:</b> {apteka_info['inn']}\n"
        message_text += f"<b>Manzil:</b> {apteka_info['adress']}\n"
        message_text += f"<b>Telefon:</b> {apteka_info['kontakt']}\n\n"
        message_text += f"<b>Foydalanuvchi:</b> {user_info.get('first_name', '')} {user_info.get('last_name', '')}\n\n"
        message_text += f"<b>Buyurtma:</b>\n"
        total_items = 0
        for i, (medicine, qty) in enumerate(order.items(), 1):
            db_medicine = MEDICINE_MAPPING.get(medicine, medicine)
            price, _ = await get_medicine_details(db_medicine, bot.pool)
            price_text = f"{price:,.0f} so'm" if price else "Narx mavjud emas"
            message_text += f"{i}. {medicine} ({db_medicine}): {qty} ta x {price_text}\n"
            total_items += qty
        message_text += f"\n<b>Jami:</b> {total_items} ta dori\n"
        message_text += f"<b>Umumiy narx:</b> {total_price:,.0f} so'm\n"
        if is_full_payment and discounted_price < total_price:
            message_text += f"<b>Chegirma bilan narx:</b> {discounted_price:,.0f} so'm\n"
            message_text += f"<b>To'lov turi:</b> 100% to'lov\n"
        else:
            message_text += f"<b>To'lov turi:</b> Chegirmasiz\n"
        await bot.send_message(chat_id=GROUP_CHAT_ID, text=message_text)
        excel_buffer = await generate_excel(order, apteka_info, total_price, discounted_price, is_full_payment)
        await bot.send_document(
            chat_id=GROUP_CHAT_ID,
            document=FSInputFile(excel_buffer, filename=f"Buyurtma_{apteka_info['inn']}_{datetime.now().strftime('%d-%m-%y_%H:%M:%S')}.xlsx"),
            caption="Buyurtma detallari Excel faylda"
        )
        logger.info(f"Buyurtma va Excel fayl guruhga yuborildi: {apteka_info['firma']}")
    except Exception as e:
        logger.error(f"Guruhga xabar yuborishda xato: {e}")

async def generate_excel(order: dict, apteka_info: dict, total_price: float, discounted_price: float, is_full_payment: bool):
    wb = Workbook()
    ws = wb.active
    ws.title = "Buyurtma"
    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['L'].width = 20
    ws['C3'] = f"Приложение к дог № {apteka_info.get('dogovor', '')}"
    ws.merge_cells('C3:L3')
    ws['C3'].font = Font(bold=True)
    ws['C3'].alignment = Alignment(horizontal='center')
    ws['C5'] = "Поставщик: MCHJ \"PERFECTFOODLAB\""
    ws.merge_cells('C5:F5')
    ws['C5'].font = Font(bold=True)
    ws['C6'] = "АДРЕС: Toshkent shaxri Chilonzor tumani. Dumbirobod 4 tor kuchasi 23/2"
    ws.merge_cells('C6:F6')
    ws['C6'].font = Font(bold=True)
    ws['C7'] = "ТЕЛ.: 71-279-85-55"
    ws.merge_cells('C7:F7')
    ws['C7'].font = Font(bold=True)
    ws['C8'] = "ИНН: 123456789"
    ws.merge_cells('C8:F8')
    ws['C8'].font = Font(bold=True)
    ws['C9'] = "Р/с: 1234 5678 9012 3456 7890 МФО:00000"
    ws.merge_cells('C9:F9')
    ws['C9'].font = Font(bold=True)
    ws['C10'] = "Регист. код плател. НДС: 456465461232165"
    ws.merge_cells('C10:F10')
    ws['C10'].font = Font(bold=True)
    ws['I5'] = "ПОКУПАТЕЛЬ:"
    ws.merge_cells('I5:J5')
    ws['K5'] = apteka_info.get('firma', '')
    ws.merge_cells('K5:L5')
    ws['I6'] = "АДРЕС:"
    ws.merge_cells('I6:J6')
    ws['K6'] = apteka_info.get('adress', '')
    ws.merge_cells('K6:L6')
    ws['I7'] = "ТЕЛ.:"
    ws.merge_cells('I7:J7')
    ws['K7'] = apteka_info.get('kontakt', '')
    ws.merge_cells('K7:L7')
    ws['I8'] = "ИНН:"
    ws.merge_cells('I8:J8')
    ws['K8'] = apteka_info.get('inn', '')
    ws.merge_cells('K8:L8')
    ws['I9'] = "Р/с"
    ws.merge_cells('I9:J9')
    ws['K9'] = apteka_info.get('rs', '')
    ws.merge_cells('K9:L9')
    ws['I10'] = "Банк МФО"
    ws.merge_cells('I10:J10')
    ws['K10'] = apteka_info.get('mfo', '')
    ws.merge_cells('K10:L10')
    ws['I5'].font = ws['K5'].font = ws['I6'].font = ws['K6'].font = ws['I7'].font = ws['K7'].font = \
    ws['I8'].font = ws['K8'].font = ws['I9'].font = ws['K9'].font = ws['I10'].font = ws['K10'].font = Font(bold=True)
    headers = ["№", "НОМЕНКЛАТУРА", "Серия", "ИКПУ", "Кол-во", "Цена", "Стоимость поставки", "НДС", "", "Стоим. Поставки с учетом НДС"]
    for col, header in enumerate(headers, start=3):
        cell = ws.cell(row=11, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    ws.merge_cells('J11:K11')
    non_discountable_medicines = {'MAGNIY B', 'KALTSIY+D3', 'ARTROZIT'}
    row = 12
    total_quantity = 0
    total_delivery = 0
    total_nds = 0
    total_with_nds = 0
    for idx, (display_medicine, qty) in enumerate(order.items(), 1):
        db_medicine = MEDICINE_MAPPING.get(display_medicine, display_medicine)
        price, _ = await get_medicine_details(db_medicine, bot.pool)
        if not price:
            price = 0
        final_price = price
        if db_medicine not in non_discountable_medicines and is_full_payment:
            if total_price > 6_000_000:
                final_price = price * (1 - 0.08)
            elif total_price > 3_000_000:
                final_price = price * (1 - 0.05)
        adjusted_price = (final_price * 100) / 112
        medicine_info = await get_medicine_info(db_medicine)
        full_name = medicine_info.get('name', display_medicine) if medicine_info else display_medicine
        ws[f'C{row}'] = idx
        ws[f'D{row}'] = full_name
        ws[f'E{row}'] = ""
        ws[f'F{row}'] = "123456887"
        ws[f'G{row}'] = qty
        ws[f'H{row}'] = round(adjusted_price, 2)
        delivery_cost = round(adjusted_price * qty, 2)
        ws[f'I{row}'] = delivery_cost
        ws[f'J{row}'] = "12"
        nds_value = round(delivery_cost * 0.12, 2)
        ws[f'K{row}'] = nds_value
        ws[f'L{row}'] = round(delivery_cost + nds_value, 2)
        total_quantity += qty
        total_delivery += delivery_cost
        total_nds += nds_value
        total_with_nds += delivery_cost + nds_value
        row += 1
    ws[f'D{row}'] = "ВСЕГО"
    ws.merge_cells(f'D{row}:F{row}')
    ws[f'G{row}'] = total_quantity
    ws[f'I{row}'] = round(total_delivery, 2)
    ws[f'K{row}'] = round(total_nds, 2)
    ws[f'L{row}'] = round(total_with_nds, 2)
    row += 3
    ws[f'D{row}'] = "ПОСТАВШИК"
    ws[f'K{row}'] = "ПОКУПАТЕЛ"
    ws.merge_cells(f'D{row}:F{row}')
    ws.merge_cells(f'K{row}:L{row}')
    ws[f'D{row}'].font = ws[f'K{row}'].font = Font(bold=True)
    ws[f'D{row}'].alignment = ws[f'K{row}'].alignment = Alignment(horizontal='center')
    row += 1
    ws[f'D{row}'] = "Директор: КАРАБАЕВ.Ф.Қ"
    ws[f'K{row}'] = "Директор: "
    ws.merge_cells(f'D{row}:F{row}')
    ws.merge_cells(f'K{row}:L{row}')
    ws[f'D{row}'].font = ws[f'K{row}'].font = Font(bold=True)
    ws[f'D{row}'].alignment = ws[f'K{row}'].alignment = Alignment(horizontal='center')
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# Bot obyektida cache'ni boshlash
bot.user_info_cache = {}

# --- Start command ---
@dp.message(CommandStart())
async def cmd_start(message: Message, state: FSMContext):
    await message.answer("Iltimos, telefon raqamingizni kiriting. (Masalan: +998901234567)")
    await state.set_state(StartWorkState.waiting_for_phone)

# --- Raqamni qabul qilish va tekshirish ---
@dp.message(StartWorkState.waiting_for_phone)
async def check_phone(message: Message, state: FSMContext):
    user_input = message.text.strip()
    digits = ''.join(filter(str.isdigit, user_input))
    if len(digits) < 9:
        await message.answer("Iltimos, to'g'ri telefon raqamini kiriting. (998901234567 yoki 901234567 ko'rinishida)")
        return
    last_9 = digits[-9:]
    try:
        async with bot.pool.acquire() as conn:
            query = """SELECT first_name, last_name FROM mp_table
            WHERE RIGHT(REGEXP_REPLACE(phone, '\D', '', 'g'), 9) = $1
            """
            row = await conn.fetchrow(query, last_9)
    except Exception as e:
        logger.error(f"DB error: {e}")
        await message.answer("Bazaga ulanishda xatolik yuz berdi.")
        return
    if row:
        bot.user_info_cache[message.from_user.id] = {
            'first_name': row['first_name'] or '',
            'last_name': row['last_name'] or '',
            'phone': last_9
        }
        await message.answer(
            f"<b>{row['first_name']} {row['last_name']}</b>, xush kelibsiz!\n\n"
            f"Iltimos, joriy geolokatsiyangizni yuboring:",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[[KeyboardButton(text="Geolokatsiya yuborish", request_location=True)]],
                resize_keyboard=True,
                one_time_keyboard=True
            )
        )
        await state.set_state(StartWorkState.waiting_for_location)
    else:
        await message.answer("Bazadan nomer topilmadi. Qayta urinib ko'ring yoki @s_saidjanov ga murojaat qiling.")

# Asosiy menyuni ko'rsatish funksiyasi
async def show_main_menu(message: Message):
    main_menu = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="Prays list")],
            [KeyboardButton(text="Spec tuzish")],
            [KeyboardButton(text="Ishni tugatish")]
        ],
        resize_keyboard=True
    )
    await message.answer("Quyidagi bo'limlardan birini tanlang:", reply_markup=main_menu)

# Geolokatsiyani qabul qilish
@dp.message(F.location)
async def handle_location(message: Message, state: FSMContext):
    latitude = message.location.latitude
    longitude = message.location.longitude
    current_state = await state.get_state()
    if current_state == StartWorkState.waiting_for_location.state:
        await state.update_data(latitude=latitude, longitude=longitude)
        await message.answer(
            f"Geolokatsiyangiz qabul qilindi:\n\n"
            f"📍 Кенглик: {latitude}\n"
            f"📍 Узунлик: {longitude}\n\n"
            f"Iltimos, kamida 10 soniyalik dumaloq video-selfi yuboring.",
            reply_markup=ReplyKeyboardRemove()
        )
        await state.set_state(StartWorkState.waiting_for_video)
    elif current_state == SpecState.selecting_search_type.state:
        await message.answer(
            f"Geolokatsiyangiz qabul qilindi:\n\n"
            f"📍 Кенглик: {latitude}\n"
            f"📍 Узунлик: {longitude}",
            reply_markup=ReplyKeyboardRemove()
        )
        search_menu = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="INN orqali qidirish")],
                [KeyboardButton(text="Nom bilan qidirish")],
                [KeyboardButton(text="🔙 Ortga")]
            ],
            resize_keyboard=True
        )
        await message.answer("Qidiruv turini tanlang:", reply_markup=search_menu)
        await state.set_state(SpecState.selecting_search_type)
    elif current_state == EndWorkState.waiting_for_location.state:
        await state.update_data(latitude=latitude, longitude=longitude)
        await message.answer(
            f"Geolokatsiyangiz qabul qilindi:\n\n"
            f"📍 Кенглик: {latitude}\n"
            f"📍 Узунлик: {longitude}\n\n"
            f"Iltimos, kamida 10 soniyalik dumaloq video-selfi yuboring.",
            reply_markup=ReplyKeyboardRemove()
        )
        await state.set_state(EndWorkState.waiting_for_video)

# Video-selfi qabul qilish
@dp.message(F.video_note)
async def handle_video_note(message: Message, state: FSMContext):
    video_note = message.video_note
    if video_note.duration < 10:
        await message.answer("Video juda qisqa. Iltimos, kamida 10 soniyalik video-selfi yuboring.")
        return
    current_state = await state.get_state()
    if current_state == StartWorkState.waiting_for_video.state:
        await message.answer(
            "Video-selfi qabul qilindi.\n\n"
            "Boshlangan ish kuningiz xayrli bo'lsin!",
            reply_markup=ReplyKeyboardRemove()
        )
        await show_main_menu(message)
        await state.clear()
    elif current_state == EndWorkState.waiting_for_video.state:
        await message.answer(
            "Video-selfi qabul qilindi.\n\n"
            "Ish kuni yakunlandi. Xayrli kun!",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[[KeyboardButton(text="Ishni boshlash")]],
                resize_keyboard=True,
                one_time_keyboard=True
            )
        )
        await state.clear()
        await state.set_state(StartWorkState.waiting_for_location)

# "Ishni boshlash" knopkasini qayta ishlash
@dp.message(F.text == "Ishni boshlash")
async def restart_work(message: Message, state: FSMContext):
    user_id = message.from_user.id
    user_info = bot.user_info_cache.get(user_id)
    if not user_info or 'phone' not in user_info:
        await message.answer(
            "Iltimos, telefon raqamingizni kiriting. (Masalan: +998901234567)",
            reply_markup=ReplyKeyboardRemove()
        )
        await state.set_state(StartWorkState.waiting_for_phone)
    else:
        await message.answer(
            f"<b>{user_info.get('first_name', '')} {user_info.get('last_name', '')}</b>, xush kelibsiz!\n\n"
            f"Iltimos, joriy geolokatsiyangizni yuboring:",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[[KeyboardButton(text="Geolokatsiya yuborish", request_location=True)]],
                resize_keyboard=True,
                one_time_keyboard=True
            )
        )
        await state.set_state(StartWorkState.waiting_for_location)

# Prays listni ko'rsatish
@dp.message(F.text == "Prays list")
async def show_price_list(message: Message):
    try:
        async with bot.pool.acquire() as conn:
            query = """SELECT image_path FROM price_list WHERE image_path IS NOT NULL ORDER BY id DESC LIMIT 1"""
            row = await conn.fetchrow(query)
    except Exception as e:
        logger.error(f"DB error: {e}")
        await message.answer("Bazadan rasm olishda xatolik yuz berdi")
        return
    if not row:
        await message.answer("Rasm topilmadi")
        return
    image_path = row["image_path"]
    try:
        await message.answer_photo(photo=FSInputFile(image_path), caption="Price list")
    except Exception as e:
        logger.error(f"File error: {e}")
        await message.answer("Rasm yo'q")

# Spec tuzish - lokatsiya so'rash
@dp.message(F.text == "Spec tuzish")
async def start_spec(message: Message, state: FSMContext):
    await message.answer(
        "Iltimos, joriy geolokatsiyangizni yuboring:",
        reply_markup=ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="Geolokatsiya yuborish", request_location=True)]],
            resize_keyboard=True,
            one_time_keyboard=True
        )
    )
    await state.set_state(SpecState.selecting_search_type)

# Qidiruv turini tanlash
@dp.message(SpecState.selecting_search_type)
async def select_search_type(message: Message, state: FSMContext):
    if message.text == "INN orqali qidirish":
        await message.answer("Iltimos, dorixonaning <b>INN</b> raqamini kiriting:")
        await state.set_state(SpecState.waiting_for_inn)
    elif message.text == "Nom bilan qidirish":
        await message.answer("Iltimos, dorixona nomini kiriting (lotincha yoki kirilchada):")
        await state.set_state(SpecState.waiting_for_name)
    elif message.text == "🔙 Ortga":
        await state.clear()
        await show_main_menu(message)

# Nom bilan qidirish
@dp.message(SpecState.waiting_for_name)
async def search_by_name(message: Message, state: FSMContext):
    search_name = message.text.strip()
    try:
        async with bot.pool.acquire() as conn:
            query = """
            SELECT inn, firma, adress, kontakt, dogovor, rs, mfo
            FROM apteka 
            WHERE firma ILIKE $1 
            ORDER BY firma
            """
            rows = await conn.fetch(query, f"%{search_name}%")
    except Exception as e:
        logger.error(f"DB error: {e}")
        await message.answer("Bazaga ulanishda xatolik yuz berdi.")
        return
    if not rows:
        await message.answer(
            "❌ Bunday nomli dorixona topilmadi.\n"
            "Boshqa nom bilan qidiring yoki INN orqali qidirish tugmasini bosing."
        )
        return
    await state.update_data(search_results=rows, current_page=0)
    await show_search_results(message, state)

# Qidiruv natijalarini sahifalab ko'rsatish funksiyasi
async def show_search_results(message: Message, state: FSMContext):
    data = await state.get_data()
    search_results = data.get('search_results', [])
    current_page = data.get('current_page', 0)
    page_size = 10
    start_idx = current_page * page_size
    end_idx = start_idx + page_size
    page_results = search_results[start_idx:end_idx]
    if not page_results:
        await message.answer("Natijalar topilmadi.")
        return
    keyboard = []
    for i, row in enumerate(page_results):
        button_text = f"{row['firma'][:30]}..." if len(row['firma']) > 30 else row['firma']
        keyboard.append([InlineKeyboardButton(
            text=button_text,
            callback_data=f"select_apteka_{start_idx + i}"
        )])
    nav_buttons = []
    if current_page > 0:
        nav_buttons.append(InlineKeyboardButton(
            text="⬅️ Oldingi",
            callback_data=f"page_{current_page - 1}"
        ))
    if end_idx < len(search_results):
        nav_buttons.append(InlineKeyboardButton(
            text="Keyingi ➡️",
            callback_data=f"page_{current_page + 1}"
        ))
    if nav_buttons:
        keyboard.append(nav_buttons)
    inline_keyboard = InlineKeyboardMarkup(inline_keyboard=keyboard)
    no_adress = "Manzil ko'rsatilmagan"
    no_phone = "Telefon ko'rsatilmagan"
    result_text = f"<b>Topilgan dorixonalar (Jami: {len(search_results)} ta, {current_page + 1}-sahifa):</b>\n\n"
    for i, row in enumerate(page_results, start_idx + 1):
        result_text += f"<b>{i}.</b> <i>{row['firma']}</i>\n"
        result_text += f"📍 {row['adress'] or no_adress}\n"
        result_text += f"📞 {row['kontakt'] or no_phone}\n"
        result_text += f"🏢 INN: {row['inn']}\n\n"
    result_text += "Kerakli dorixonani tanlang:"
    if message.text:
        await message.answer(result_text, reply_markup=inline_keyboard)
    else:
        await message.edit_text(result_text, reply_markup=inline_keyboard)

# Callback handler for apteka selection
@dp.callback_query(F.data.startswith("select_apteka_"))
async def select_apteka_from_list(callback_query, state: FSMContext):
    await callback_query.answer()
    try:
        selected_index = int(callback_query.data.split("_")[-1])
        data = await state.get_data()
        search_results = data.get('search_results', [])
        if selected_index < len(search_results):
            selected_apteka = search_results[selected_index]
            no_adress = "Manzil ko'rsatilmagan"
            no_phone = "Telefon ko'rsatilmagan"
            await state.update_data(
                inn=selected_apteka['inn'],
                firma=selected_apteka['firma'],
                adress=selected_apteka['adress'],
                contact=selected_apteka['kontakt'],
                dogovor=selected_apteka['dogovor'],
                rs=selected_apteka['rs'],
                mfo=selected_apteka['mfo']
            )
            confirm_text = f"<b>Tanlangan dorixona:</b>\n\n"
            confirm_text += f"<b>Nomi:</b> {selected_apteka['firma']}\n"
            confirm_text += f"<b>INN:</b> {selected_apteka['inn']}\n"
            confirm_text += f"<b>Manzil:</b> {selected_apteka['adress'] or no_adress}\n"
            confirm_text += f"<b>Telefon:</b> {selected_apteka['kontakt'] or no_phone}\n\n"
            confirm_text += "Shu dorixona ekanligini tasdiqlaysizmi?"
            buttons = ReplyKeyboardMarkup(
                keyboard=[[KeyboardButton(text="Ha✅"), KeyboardButton(text="Yo'q❌")]],
                resize_keyboard=True,
                one_time_keyboard=True
            )
            await callback_query.message.edit_reply_markup(reply_markup=None)
            await callback_query.message.answer(confirm_text, reply_markup=buttons)
            await state.set_state(SpecState.confirming_apteka)
        else:
            await callback_query.message.answer("Xatolik yuz berdi. Qaytadan urinib ko'ring.")
    except Exception as e:
        logger.error(f"Callback error: {e}")
        await callback_query.message.answer("Xatolik yuz berdi. Qaytadan urinib ko'ring.")

# Callback handler for pagination
@dp.callback_query(F.data.startswith("page_"))
async def handle_pagination(callback_query, state: FSMContext):
    await callback_query.answer()
    try:
        new_page = int(callback_query.data.split("_")[-1])
        await state.update_data(current_page=new_page)
        await show_search_results(callback_query.message, state)
    except Exception as e:
        logger.error(f"Pagination error: {e}")
        await callback_query.message.answer("Xatolik yuz berdi. Qaytadan urinib ko'ring.")

# INN raqamini tekshirish
@dp.message(SpecState.waiting_for_inn)
async def check_inn(message: Message, state: FSMContext):
    if message.text == "Nom bilan qidirish":
        await message.answer("Iltimos, dorixona nomini kiriting (lotincha yoki kirilchada):")
        await state.set_state(SpecState.waiting_for_name)
        return
    elif message.text == "🔙 Ortga":
        search_menu = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="INN orqali qidirish")],
                [KeyboardButton(text="Nom bilan qidirish")],
                [KeyboardButton(text="🔙 Ortga")]
            ],
            resize_keyboard=True
        )
        await message.answer("Qidiruv turini tanlang:", reply_markup=search_menu)
        await state.set_state(SpecState.selecting_search_type)
        return
    inn_input = message.text.strip()
    try:
        async with bot.pool.acquire() as conn:
            query = "SELECT firma, adress, kontakt, dogovor, rs, mfo FROM apteka WHERE inn = $1"
            row = await conn.fetchrow(query, inn_input)
    except Exception as e:
        logger.error(f"DB error: {e}")
        await message.answer("Bazaga ulanishda xatolik yuz berdi.")
        return
    if row:
        await state.update_data(
            inn=inn_input,
            firma=row["firma"],
            adress=row["adress"],
            contact=row["kontakt"],
            dogovor=row["dogovor"],
            rs=row["rs"],
            mfo=row["mfo"]
        )
        buttons = ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="Ha✅"), KeyboardButton(text="Yo'q❌")]],
            resize_keyboard=True,
            one_time_keyboard=True
        )
        no_adress = "Manzil ko'rsatilmagan"
        no_phone = "Telefon ko'rsatilmagan"
        confirm_text = f"Siz kiritgan INN bazadan topildi:\n\n"
        confirm_text += f"<b>Dorixona nomi:</b> <i>{row['firma']}</i>\n"
        confirm_text += f"<b>INN:</b> {inn_input}\n"
        confirm_text += f"<b>Manzil:</b> {row['adress'] or no_adress}\n"
        confirm_text += f"<b>Telefon:</b> {row['kontakt'] or no_phone}\n\n"
        confirm_text += "Shu dorixona ekanligini tasdiqlaysizmi?"
        await message.answer(confirm_text, reply_markup=buttons)
        await state.set_state(SpecState.confirming_apteka)
    else:
        await message.answer(
            "❌ Bu apteka bazada yo'q.\n"
            "Iltimos, INN raqamini to'g'ri kiritganingizga ishonch hosil qiling yoki adminlarga murojaat qiling."
        )
        await message.answer("Iltimos, dorixonaning <b>INN</b> raqamini kiriting:")

# Aptekani tasdiqlash
@dp.message(SpecState.confirming_apteka)
async def confirm_apteka(message: Message, state: FSMContext):
    if message.text == "Ha✅":
        data = await state.get_data()
        display_medicines = list(MEDICINE_MAPPING.keys())
        if not display_medicines:
            await message.answer("Dorilar ro'yxati bo'sh.")
            await state.clear()
            await show_main_menu(message)
            return
        await state.update_data(order={})
        await state.set_state(OrderState.selecting_medicine)
        await show_medicines_list(message, state)
    elif message.text == "Yo'q❌":
        await message.answer("Qaytadan qidiring.")
        search_menu = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="INN orqali qidirish")],
                [KeyboardButton(text="Nom bilan qidirish")],
                [KeyboardButton(text="🔙 Ortga")]
            ],
            resize_keyboard=True
        )
        await message.answer("Qidiruv turini tanlang:", reply_markup=search_menu)
        await state.set_state(SpecState.selecting_search_type)

# Dori ma'lumotlarini olish funksiyasi
async def get_medicine_info(db_medicine_name):
    return DORIOPS.get(db_medicine_name, None)

async def show_medicines_list(message: Message, state: FSMContext):
    data = await state.get_data()
    order = data.get('order', {})
    display_medicines = list(MEDICINE_MAPPING.keys())
    if not display_medicines:
        await message.answer("Dorilar ro'yxati bo'sh.")
        return
    keyboard = []
    for i in range(0, len(display_medicines), 2):
        row = []
        for j in range(i, min(i + 2, len(display_medicines))):
            medicine = display_medicines[j]
            if medicine in order:
                row.append(KeyboardButton(text=f"{medicine} ({order[medicine]})"))
            else:
                row.append(KeyboardButton(text=medicine))
        keyboard.append(row)
    if order:
        keyboard.append([KeyboardButton(text="Spec yakunlash")])
    keyboard.append([KeyboardButton(text="🔙 Ortga")])
    medicines_keyboard = ReplyKeyboardMarkup(keyboard=keyboard, resize_keyboard=True)
    firma = data.get('firma', 'Dorixona')
    order_summary = f"<b>Dorixona:</b> {firma}\n"
    if order:
        order_summary += f"\n<b>Tanlangan dorilar ({len(order)} ta):</b>\n"
        for med, qty in order.items():
            order_summary += f"• {med}: {qty} ta\n"
    await message.answer(f"{order_summary}\nKerakli dorilarni tanlang:", reply_markup=medicines_keyboard)

# Dori tanlash
@dp.message(OrderState.selecting_medicine)
async def select_medicine(message: Message, state: FSMContext):
    if message.text == "🔙 Ortga":
        await state.clear()
        await show_main_menu(message)
        return
    elif message.text == "Spec yakunlash":
        data = await state.get_data()
        order = data.get('order', {})
        order = {k: v for k, v in order.items() if v > 0}
        await state.update_data(order=order)
        if not order:
            await message.answer("Hech qanday dori tanlanmagan. Iltimos, kamida bitta dori tanlang.")
            return
        payment_buttons = ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="Ha✅"), KeyboardButton(text="Yo'q❌")]],
            resize_keyboard=True,
            one_time_keyboard=True
        )
        await message.answer("To'lov 100% bo'ladimi?", reply_markup=payment_buttons)
        await state.set_state(OrderState.reviewing_order)
        return
    selected_medicine = message.text.strip()
    if '(' in selected_medicine:
        selected_medicine = selected_medicine.split('(')[0].strip()
    if selected_medicine not in MEDICINE_MAPPING:
        await message.answer("Iltimos, ro'yxatdan dorini tanlang!")
        await show_medicines_list(message, state)
        return
    db_medicine = MEDICINE_MAPPING.get(selected_medicine, selected_medicine)
    medicine_info = await get_medicine_info(db_medicine)
    if medicine_info:
        info_text = f"<b>{selected_medicine}</b>\n"
        info_text += f"<b>Tavsif:</b> {medicine_info.get('tavsif')}\n"
        info_text += f"<b>Hajm:</b> {medicine_info.get('hajm')}\n"
        rasm_path = medicine_info.get('rasm_path')
        if rasm_path and os.path.exists(rasm_path):
            try:
                await message.answer_photo(photo=FSInputFile(rasm_path), caption=info_text)
            except Exception as e:
                logger.error(f"Rasm yuborishda xato, dori: {selected_medicine}, xato: {e}")
                await message.answer(info_text)
                await message.answer("Rasmni yuborishda xatolik yuz berdi.")
        else:
            await message.answer(info_text)
            if rasm_path:
                logger.warning(f"Rasm fayli topilmadi: {rasm_path}")
                await message.answer("Rasm fayli mavjud emas.")
    else:
        await message.answer(f"<b>{selected_medicine}</b> haqida ma'lumot topilmadi.")
    data = await state.get_data()
    order = data.get('order', {})
    if selected_medicine in order:
        current_qty = order[selected_medicine]
        edit_keyboard = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="Miqdorni o'zgartirish")],
                [KeyboardButton(text="Dorini o'chirish")],
                [KeyboardButton(text="🔙 Orqaga qaytish")]
            ],
            resize_keyboard=True,
            one_time_keyboard=True
        )
        await message.answer(
            f"<b>{selected_medicine}</b> allaqachon buyurtmada mavjud!\n\n"
            f"<b>Joriy miqdor:</b> {current_qty} ta\n\n"
            f"Nima qilishni istaysiz?",
            reply_markup=edit_keyboard
        )
        await state.update_data(selected_medicine=selected_medicine, editing_mode=True)
        await state.set_state(OrderState.entering_quantity)
    else:
        await message.answer(
            f"<b>{selected_medicine}</b> uchun miqdor kiriting (raqam bilan, masalan: 5):",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[[KeyboardButton(text="🔙 Orqaga qayt")]],
                resize_keyboard=True
            )
        )
        await state.update_data(selected_medicine=selected_medicine, editing_mode=False)
        await state.set_state(OrderState.entering_quantity)

@dp.message(OrderState.entering_quantity)
async def enter_quantity(message: Message, state: FSMContext):
    if message.text == "🔙 Orqaga qaytish":
        await state.set_state(OrderState.selecting_medicine)
        await show_medicines_list(message, state)
        return
    data = await state.get_data()
    selected_medicine = data.get('selected_medicine')
    editing_mode = data.get('editing_mode', False)
    order = data.get('order', {})
    if editing_mode:
        if message.text == "Miqdorni o'zgartirish":
            current_qty = order.get(selected_medicine, 0)
            await message.answer(
                f"<b>{selected_medicine}</b> uchun yangi miqdor kiriting:\n\n"
                f"<b>Joriy miqdor:</b> {current_qty} ta\n\n"
                f"Yangi miqdor (0 kiritilsa dori o'chiriladi):",
                reply_markup=ReplyKeyboardMarkup(
                    keyboard=[[KeyboardButton(text="🔙 Orqaga qaytish")]],
                    resize_keyboard=True
                )
            )
            await state.update_data(editing_mode=False)
            return
        elif message.text == "Dorini o'chirish":
            if selected_medicine in order:
                del order[selected_medicine]
                await message.answer(
                    f"<b>{selected_medicine}</b> buyurtmadan o'chirildi!",
                    reply_markup=ReplyKeyboardRemove()
                )
            else:
                await message.answer(
                    f"<b>{selected_medicine}</b> buyurtmada yo'q edi.",
                    reply_markup=ReplyKeyboardRemove()
                )
            await state.update_data(order=order)
            await state.set_state(OrderState.selecting_medicine)
            await show_medicines_list(message, state)
            return
    try:
        quantity = int(message.text.strip())
        if quantity < 0:
            await message.answer("Iltimos, manfiy bo'lmagan raqam kiriting.")
            return
    except ValueError:
        await message.answer("Iltimos, faqat raqam kiriting (masalan: 5).")
        return
    if quantity == 0:
        if selected_medicine in order:
            del order[selected_medicine]
            await message.answer(
                f"<b>{selected_medicine}</b> buyurtmadan o'chirildi.\n"
                f"Yana dori tanlash uchun davom eting.",
                reply_markup=ReplyKeyboardRemove()
            )
        else:
            await message.answer(
                f"<b>{selected_medicine}</b> buyurtmada yo'q edi.\n"
                f"Yana dori tanlash uchun davom eting.",
                reply_markup=ReplyKeyboardRemove()
            )
    else:
        old_qty = order.get(selected_medicine, 0)
        order[selected_medicine] = quantity
        if old_qty > 0:
            await message.answer(
                f"<b>{selected_medicine}</b>: {old_qty} ta → {quantity} ta (yangilandi)\n"
                f"Yana dori tanlash uchun davom eting.",
                reply_markup=ReplyKeyboardRemove()
            )
        else:
            await message.answer(
                f"<b>{selected_medicine}</b>: {quantity} ta qo'shildi.\n"
                f"Yana dori tanlash uchun davom eting.",
                reply_markup=ReplyKeyboardRemove()
            )
    await state.update_data(order=order)
    await state.set_state(OrderState.selecting_medicine)
    await show_medicines_list(message, state)

# Ishni tugatish
@dp.message(F.text == "Ishni tugatish")
async def end_work(message: Message, state: FSMContext):
    await message.answer(
        "Iltimos, joriy geolokatsiyangizni yuboring:",
        reply_markup=ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="Geolokatsiya yuborish", request_location=True)]],
            resize_keyboard=True,
            one_time_keyboard=True
        )
    )
    await state.set_state(EndWorkState.waiting_for_location)

if __name__ == "__main__":
    async def main():
        pool = await asyncpg.create_pool(**DB_CONFIG, min_size=1, max_size=10)
        bot.pool = pool
        try:
            await dp.start_polling(bot)
        finally:
            await pool.close()
    asyncio.run(main())